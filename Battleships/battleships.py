#*************************
# Battleship game using Excel and Pandas (and openpyxl)
# Created by: Huang ChenTing
# Contact: huang.chenting(AT)gmail.com
# Date: Jun 2023
#*************************

import pandas as pd
import openpyxl
import random

MAPSIZE = 25
SHIPNUM = 30
SHIPMAXLEN = 10

# a bunch of color options to make it easier to see in excel
darkerRedFill = openpyxl.styles.PatternFill(fgColor='88880000', fill_type='solid')
fillColor = [openpyxl.styles.PatternFill(fgColor='FFFF0000', fill_type='solid'),
             openpyxl.styles.PatternFill(fgColor='FFFFFF00', fill_type='solid'),
             openpyxl.styles.PatternFill(fgColor='FF00FFFF', fill_type='solid'),
             openpyxl.styles.PatternFill(fgColor='000FFF00', fill_type='solid'),
             openpyxl.styles.PatternFill(fgColor='FF0000FF', fill_type='solid')]

used_coords = []    # to track no overlapping of ships during generation
ship_list = []      # to keep track of how many ships sunk later


# this function generate the positions of ships on a grid. 
# ships can be of random length, randomly vertical or horizontal
# but ships cannot overlap other ships
def GenerateShipPositions():
    used_coords.clear()
    ship_list.clear()
    
    for i in range(SHIPNUM): # limit number of ships to create
        
        ship_pos = []
        ship_placed = False
                
        while ship_placed == False:
            ship_len = random.randint(2, SHIPMAXLEN)
            vert = random.choice([True, False])
            start_pt = (random.randint(1, MAPSIZE - ship_len + 1), random.randint(1, MAPSIZE - ship_len + 1))
                    
            # if vertical
            if vert:
                for j in range(ship_len):
                    ship_pos.append((start_pt[0] + j, start_pt[1]))
            else: # horizontal
                for j in range(ship_len):
                    ship_pos.append((start_pt[0], start_pt[1] + j))

            # assumed placed correctly
            ship_placed = True
            
            # but check if any point is in used coords
            for p in ship_pos:
                if p in used_coords:
                    ship_placed = False # false
                    ship_pos.clear()    # clear, should exit for loop immediately
                    
            # if still true
            if ship_placed:
                for p in ship_pos:
                    used_coords.append(p)
                    
        #reached here means ship is placed, so append to a ship list to keep track
        ship_list.append(ship_pos)
            
            
# use openpyxl to create an excel file, to place our created ships onto the 'map'
def GenerateBattleShipMap():         

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Battleships Map"
    
    # fill up map with zeroes first
    for i in range(MAPSIZE):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 3.2 # magic number, does not correspond to excel column hmmm
        for j in range(MAPSIZE):
            ws.cell(row = i + 1, column = j + 1).value = 0
       
    # mark down ship positions on map
    index = 0
    for s in ship_list:
        # print(s) for debug
        ship_color = fillColor[index % 5]
        index += 1
        for s_pos in s:
            #ws.cell (row = s_pos[0], column = s_pos[1]).value = str(s_pos[0]) + ", " + str(s_pos[1]) for debug
            ws.cell (row = s_pos[0], column = s_pos[1]).value = 1
            ws.cell (row = s_pos[0], column = s_pos[1]).fill = ship_color
            
    # save, so we can reference it later
    wb.save('battleship_map.xlsx')
    
    return ws
                    
# compare data with our ship list, to find out whether our ships have been sunk
def CheckShipStatus(results_ws):
    ships_sunk = 0
    
    for s in ship_list:
        assumed_sunk = True
        for s_pos in s:
            if results_ws.cell(row = s_pos[0], column = s_pos[1]).value == "O":
                assumed_sunk = False
                continue
        
        if assumed_sunk:
            for s_pos in s:
                results_ws.cell(row = s_pos[0], column = s_pos[1]).fill = darkerRedFill
            ships_sunk += 1
            
    return ships_sunk
                

 
      

def main():

# read attacker file first see how
    print("Welcome to Excel Battleship Command")
    print("Let the battle begin!")
            
    # generate ship positions
    GenerateShipPositions()
    
    # use openpyxl to create a battleship map
    battleship_map = GenerateBattleShipMap()
       
    # use pandas to read the attacker map, 
    # and then compare with battleship map which we already have
    xl = pd.ExcelFile("attacker_map.xlsm")
    df = xl.parse("Map", header = None)
    
    # open up a new file to compare
    compare_wb = openpyxl.load_workbook('battleship_map.xlsx')
    compare_ws = compare_wb.create_sheet('Battle Results')
    compare_ws.title = "Battle Results"
    
    
    for i in range(MAPSIZE):
        compare_ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 3.2 # magic number, does not correspond to excel column hmmm
        for j in range(MAPSIZE):
            if df[j][i] == battleship_map.cell(row = i + 1, column = j + 1).value == 1: # have to be equal, have to be 1 too, not == 0!
                # hit
                compare_ws.cell(row = i+1, column = j+1).value = "X"
                compare_ws.cell(row = i+1, column = j+1).fill = fillColor[0]
            else:
                compare_ws.cell(row = i+1, column = j+1).value = "O"
    
    # check for sunk ships
    ships_sunk_count = CheckShipStatus(compare_ws)
    
    if ships_sunk_count == 1:
        print("After checking, " + str(ships_sunk_count) + " ship was sunk")
    else:
        print("After checking, " + str(ships_sunk_count) + " ships were sunk")
    print("Open up battleship_map file, Battle Results tab to verify.")
    print("Thank you for playing ~")
    
    compare_wb.save('battleship_map.xlsx')                        



if __name__ == '__main__':
    main()



